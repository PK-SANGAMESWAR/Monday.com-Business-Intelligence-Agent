"""Faithful xlsx -> rows reader for the two source workbooks.

This module's only job is transport: read every data row exactly as openpyxl
returns it, with native Python types intact, and hand back rows plus the two
kinds of row that need special handling before they can be written:

* **fully-empty rows** are dropped (a row with nothing in it is not a record
  of anything, and monday would reject or mangle it) but every drop is named,
  never silent;
* **embedded junk header rows** (Deals, Excel rows 52 and 181) are kept and
  flagged, per CLAUDE.md and F03 section 3.5 — the board must carry the same
  mess the workbook does, not a cleaned-up version of it.

pandas is deliberately not used here (F03 section 3.9): it coerces on read
(ints become floats, empty cells become NaN), which is a small unfaithfulness
this module's entire job is to avoid. `openpyxl(data_only=True)` returns
native types — `datetime`, `int`, `float`, `str`, `None` — unchanged.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from scripts.seeding.errors import WorkbookError

__all__ = [
    "DEALS_HEADER_ROW",
    "DEALS_SHEET_NAME",
    "DEALS_WORKBOOK_FILENAME",
    "JUNK_ROW_MARKER_HEADER",
    "JUNK_ROW_MARKER_VALUE",
    "WORK_ORDERS_HEADER_ROW",
    "WORK_ORDERS_SHEET_NAME",
    "WORK_ORDERS_WORKBOOK_FILENAME",
    "WorkbookReadResult",
    "WorkbookRow",
    "read_deals_workbook",
    "read_work_orders_workbook",
]

DEALS_WORKBOOK_FILENAME = "Deal funnel Data.xlsx"
DEALS_SHEET_NAME = "Deal tracker"
DEALS_HEADER_ROW = 1

WORK_ORDERS_WORKBOOK_FILENAME = "Work_Order_Tracker Data.xlsx"
WORK_ORDERS_SHEET_NAME = "work order tracker"
WORK_ORDERS_HEADER_ROW = 2

#: Embedded junk rows carry literal header text in every column except the
#: first. Detecting on `Deal Name == 'Deal Name'` misses them (column A holds
#: a real name, `Nezuko` / `Bugs Bunny`) — CLAUDE.md and the data profile both
#: point at `Deal Status` as the reliable marker.
JUNK_ROW_MARKER_HEADER = "Deal Status"
JUNK_ROW_MARKER_VALUE = "Deal Status"


@dataclass(frozen=True)
class WorkbookRow:
    """One data row, keyed by workbook header, with native Python values."""

    source_row: int
    values: dict[str, Any]
    is_junk: bool = False


@dataclass(frozen=True)
class WorkbookReadResult:
    sheet_name: str
    headers: list[str]
    rows: list[WorkbookRow]
    dropped_empty_rows: list[int]
    junk_rows: list[int]

    @property
    def row_count(self) -> int:
        return len(self.rows)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    return isinstance(value, str) and value.strip() == ""


def _read_sheet(path: Path, sheet_name: str, header_row: int) -> WorkbookReadResult:
    if not path.exists():
        raise WorkbookError(f"workbook not found: {path}")

    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WorkbookError(
                f"sheet {sheet_name!r} not found in {path.name}; "
                f"available sheets: {', '.join(workbook.sheetnames)}"
            )
        sheet = workbook[sheet_name]

        header_cells = sheet[header_row]
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in header_cells
        ]
        if not any(headers):
            raise WorkbookError(
                f"header row {header_row} of {sheet_name!r} in {path.name} is "
                "empty. Reading the wrong header row would silently produce "
                "all-None columns rather than failing loudly."
            )

        rows: list[WorkbookRow] = []
        dropped_empty: list[int] = []
        junk_rows: list[int] = []

        for row_cells in sheet.iter_rows(min_row=header_row + 1):
            row_number = row_cells[0].row
            values = {
                header: cell.value
                for header, cell in zip(headers, row_cells, strict=False)
                if header
            }
            if all(_is_blank(value) for value in values.values()):
                dropped_empty.append(row_number)
                continue

            is_junk = (
                str(values.get(JUNK_ROW_MARKER_HEADER, "")).strip()
                == JUNK_ROW_MARKER_VALUE
            )
            if is_junk:
                junk_rows.append(row_number)

            rows.append(
                WorkbookRow(source_row=row_number, values=values, is_junk=is_junk)
            )
    finally:
        workbook.close()

    return WorkbookReadResult(
        sheet_name=sheet_name,
        headers=headers,
        rows=rows,
        dropped_empty_rows=dropped_empty,
        junk_rows=junk_rows,
    )


def read_deals_workbook(path: Path) -> WorkbookReadResult:
    """Read `Deal tracker`: header on row 1, junk rows flagged not dropped."""
    return _read_sheet(path, DEALS_SHEET_NAME, DEALS_HEADER_ROW)


def read_work_orders_workbook(path: Path) -> WorkbookReadResult:
    """Read `work order tracker`: header on row 2 — row 1 is blank."""
    return _read_sheet(path, WORK_ORDERS_SHEET_NAME, WORK_ORDERS_HEADER_ROW)
